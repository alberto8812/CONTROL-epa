"""
reporter.py — Report generation, Excel writing, and OneDrive upload.

Pure functions (generate_password, build_report_filename, build_report_rows,
write_excel) have no side effects and are fully unit-testable.

Playwright-bound functions (collect_subfolders, upload_report, run_report)
require an authenticated Playwright page and are integration-tested manually.
"""

import os
import secrets
import tempfile
import urllib.error
import urllib.parse
import urllib.request
from dataclasses import dataclass, field as dc_field
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any

from loguru import logger

from onedrive_rpa import config


class ReportError(Exception):
    """Raised when a report operation fails (e.g. folder not found, upload error)."""


# ---------------------------------------------------------------------------
# T-06: generate_password
# ---------------------------------------------------------------------------


def generate_password(length: int = config.REPORT_PASSWORD_LENGTH) -> str:
    """
    Generate a cryptographically random password.

    Args:
        length: Number of characters. Must be >= 16.

    Returns:
        A random string of ``length`` characters drawn from
        ``config.REPORT_PASSWORD_ALPHABET``.

    Raises:
        ValueError: if ``length`` < 16.
    """
    if length < 16:
        raise ValueError(f"Password length must be >= 16, got {length}")
    return "".join(secrets.choice(config.REPORT_PASSWORD_ALPHABET) for _ in range(length))


# ---------------------------------------------------------------------------
# T-07: build_report_filename
# ---------------------------------------------------------------------------


def build_report_filename(now: datetime | None = None) -> str:
    """
    Build the report filename using the configured prefix and timestamp format.

    Args:
        now: Optional datetime to use as the timestamp. Defaults to datetime.now().

    Returns:
        A filename string like ``reporte_20260522_143015.xlsx``.
    """
    ts = (now or datetime.now()).strftime(config.REPORT_FILENAME_TIMESTAMP_FORMAT)
    return f"{config.REPORT_FILENAME_PREFIX}_{ts}.xlsx"


# ---------------------------------------------------------------------------
# URL assembly helper (EU-5)
# ---------------------------------------------------------------------------


def _build_folder_url(source_folder: str, name: str) -> str:
    """
    Construct the full OneDrive URL for a subfolder entry.

    Assembles the URL from config constants + per-row inputs.  Each path
    segment is stripped of surrounding slashes and URL-encoded individually so
    that spaces and special characters in folder names are handled correctly.
    Empty segments (e.g. when ``source_folder`` is ``""``) are filtered out so
    no double-slash appears in the path.

    Args:
        source_folder: The parent folder path relative to the Documents library
                       (e.g. ``"clientes"`` or ``"clientes/"``).
        name:          The subfolder name (e.g. ``"AlphaClient"``).

    Returns:
        A fully qualified URL string suitable for Fernet encryption.
    """
    base = config.ONEDRIVE_URL.rstrip("/")
    segments = [
        config.SHAREPOINT_PERSONAL_PATH,
        "Documents",
        source_folder,
        name,
    ]
    encoded_segments = [
        urllib.parse.quote(seg.strip("/"), safe="/")
        for seg in segments
        if seg.strip("/")
    ]
    return base + "/" + "/".join(encoded_segments)


def _build_folder_url_from_path(full_path: str) -> str:
    """Like _build_folder_url but accepts a single pre-assembled relative path."""
    base = config.ONEDRIVE_URL.rstrip("/")
    segments = [config.SHAREPOINT_PERSONAL_PATH, "Documents", full_path]
    encoded_segments = [
        urllib.parse.quote(seg.strip("/"), safe="/")
        for seg in segments
        if seg.strip("/")
    ]
    return base + "/" + "/".join(encoded_segments)


def _shorten_url(long_url: str) -> str:
    """Call the configured URL shortener API and return the short URL.

    Supports GET providers (is.gd, TinyURL) and POST providers (Rebrandly).
    All behaviour is controlled via .env — no code changes to switch providers.

    .env examples:
        # is.gd (GET, no key)
        URL_SHORTENER_ENDPOINT=https://is.gd/create.php?format=simple
        URL_SHORTENER_METHOD=GET

        # Rebrandly (POST, requires API key)
        URL_SHORTENER_ENDPOINT=https://api.rebrandly.com/v1/links
        URL_SHORTENER_METHOD=POST
        URL_SHORTENER_API_KEY=<your-key>
        URL_SHORTENER_KEY_HEADER=apikey
        URL_SHORTENER_BODY_KEY=destination
        URL_SHORTENER_RESPONSE_KEY=shortUrl
        URL_SHORTENER_DOMAIN=rebrand.ly   # optional

    Fail-open: any error returns the original long_url.
    """
    import json as _json
    import ssl

    endpoint = config.URL_SHORTENER_ENDPOINT
    if not endpoint:
        return long_url

    try:
        # Use certifi's CA bundle — fixes SSL errors on macOS where Python's
        # urllib does not pick up the system certificate store by default.
        try:
            import certifi
            ssl_ctx = ssl.create_default_context(cafile=certifi.where())
        except ImportError:
            ssl_ctx = ssl.create_default_context()

        headers: dict[str, str] = {
            "User-Agent": "novahold-rpa/1.0",
            "Content-Type": "application/json",
        }
        if config.URL_SHORTENER_API_KEY:
            headers[config.URL_SHORTENER_KEY_HEADER or "Authorization"] = config.URL_SHORTENER_API_KEY

        if config.URL_SHORTENER_METHOD == "POST":
            body: dict = {config.URL_SHORTENER_BODY_KEY or "destination": long_url}
            if config.URL_SHORTENER_DOMAIN:
                body["domain"] = {"fullName": config.URL_SHORTENER_DOMAIN}
            data = _json.dumps(body).encode("utf-8")
            req = urllib.request.Request(endpoint, data=data, headers=headers, method="POST")
        else:
            sep = "&" if "?" in endpoint else "?"
            api_url = f"{endpoint}{sep}url={urllib.parse.quote(long_url, safe='')}"
            req = urllib.request.Request(api_url, headers=headers)

        with urllib.request.urlopen(req, timeout=10, context=ssl_ctx) as resp:
            raw = resp.read().decode("utf-8").strip()

        response_key = config.URL_SHORTENER_RESPONSE_KEY
        if response_key:
            short = str(_json.loads(raw).get(response_key, ""))
        else:
            short = raw

        if short.startswith("http"):
            return short
        # Some providers (e.g. Rebrandly) return the short URL without scheme.
        if short and "." in short and not short.startswith("{"):
            return f"https://{short}"
        logger.warning("URL_SHORTENER | unexpected response={r} | using full URL", r=short[:80])
        return long_url

    except urllib.error.HTTPError as exc:
        # Capture response body for richer diagnostics (e.g. Rebrandly quota exceeded)
        try:
            body = exc.read().decode("utf-8")[:300]
        except Exception:
            body = ""
        logger.warning(
            "URL_SHORTENER_ERROR | status={s} | reason={r} | detail={d} | using full URL",
            s=exc.code,
            r=exc.reason,
            d=body,
        )
        return long_url
    except Exception as exc:
        logger.warning("URL_SHORTENER_ERROR | reason={r} | using full URL", r=str(exc))
        return long_url


# ---------------------------------------------------------------------------
# T-08: build_report_rows
# ---------------------------------------------------------------------------


def build_report_rows(
    folder_names: list[str],
    *,
    now: datetime | None = None,
    source_folder: str = "",
    fernet=None,
    passwords: dict[str, str] | None = None,
    share_urls: dict[str, str] | None = None,
    share_statuses: dict[str, str] | None = None,
    full_paths: list[str] | None = None,
    expiry_date: datetime | None = None,
) -> list[dict[str, Any]]:
    """
    Build the list of report row dicts from a list of folder names.

    Each row contains:
        - folder_name: the folder name as-is
        - password: pre-generated password from *passwords* map when available,
          otherwise a freshly generated random password
        - encrypted_url: Fernet-encrypted OneDrive URL (str), or ``""`` when no
          key is configured (fail-open, EU-1)
        - creation_date: the ``now`` datetime (or datetime.now())

    Args:
        folder_names: Ordered list of folder names to include in the report.
        now: Optional datetime to stamp as the creation date. Defaults to
             datetime.now(). Injecting a fixed value makes tests deterministic.
        source_folder: Parent folder path relative to Documents (e.g.
                       ``"clientes"``). Used to build the per-row OneDrive URL
                       when ``full_paths`` is not provided.
        fernet: Optional :class:`~cryptography.fernet.Fernet` instance for
                encryption.  When ``None``, falls back to ``config.FERNET``.
                Injecting a value here makes tests independent of the
                environment (EU-2).
        passwords: Optional pre-generated password map keyed by folder base name.
                   When a key matching the folder name is found, that value is
                   used instead of calling ``generate_password()``.  When
                   ``None`` or the key is absent, falls back to
                   ``generate_password()`` (backward compatible).
        share_urls: Optional map of folder base name -> OneDrive sharing URL
                    captured from the "Copiar vínculo" button after sharing.
                    When provided, the sharing URL is used as-is instead of the
                    path-based URL built by ``_build_folder_url()``.  When
                    ``None`` or a folder has no entry, falls back to the
                     constructed URL (backward compatible, fail-open).
        share_statuses: Optional map of folder base name -> ``Shared``,
                        ``Failed``, ``Skipped``, or ``Not shared``. Non-shared
                        statuses deliberately produce blank URL and password fields.
        full_paths: Optional list of full folder paths parallel to
                    ``folder_names`` (e.g. ``["Camion/Plano/Bz8yy", ...]``).
                    When provided, the fallback URL is built from the full path
                    instead of ``source_folder + name``, so the URL points to
                    the specific leaf folder rather than the container.

    Returns:
        A list of dicts, one per folder name.
    """
    creation_date = now or datetime.now()
    active_fernet = fernet if fernet is not None else config.FERNET
    rows = []
    for i, name in enumerate(folder_names):
        share_status = share_statuses.get(name, "Not shared") if share_statuses is not None else "Shared"
        if share_status == "Shared":
            # Use the real OneDrive sharing URL when available (captured from
            # "Copiar vínculo" after sharing); fall back to the constructed path URL.
            full_path = full_paths[i] if full_paths and i < len(full_paths) else None
            url = (share_urls.get(name) if share_urls else None) or (
                _build_folder_url_from_path(full_path) if full_path
                else _build_folder_url(source_folder, name)
            )
            # Short URL via configured provider (fail-open: falls back to full URL)
            short_url = _shorten_url(url)
            # Fernet token kept for auditability — lets you recover the original URL
            # from the token even if the shortener link expires.
            encrypted_url = (
                active_fernet.encrypt(url.encode()).decode("ascii")
                if active_fernet is not None else ""
            )
            pwd = passwords.get(name) if passwords else None
        else:
            # A failed or intentionally skipped share has no usable password-protected
            # link. Keep every member of that pair blank rather than implying otherwise.
            url = ""
            short_url = ""
            encrypted_url = ""
            pwd = ""
        rows.append(
            {
                "folder_name": name,
                "share_status": share_status,
                "password": pwd or (generate_password() if share_status == "Shared" else ""),
                "short_url": short_url,
                "encrypted_url": encrypted_url,
                "folder_url": url,
                "creation_date": creation_date,
                "expiry_date": expiry_date,
            }
        )
    return rows


# ---------------------------------------------------------------------------
# T-09: write_excel
# ---------------------------------------------------------------------------


def write_excel(rows: list[dict[str, Any]]) -> BytesIO:
    """
    Serialize a list of report rows to an Excel workbook in memory.

    The workbook has a single sheet titled "Report" with a bold header row
    followed by one data row per entry in ``rows``.

    Column order: folder name, share status, usable URL, encrypted URL,
    password, creation date, expiry date.

    Args:
        rows: List of dicts as produced by :func:`build_report_rows`.

    Returns:
        A :class:`~io.BytesIO` seeked to position 0, ready to be read or
        passed directly to ``openpyxl.load_workbook()``.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    headers = [
        "Folder Name", "Share Status", "URL", "Encrypted URL", "Password",
        "Creation Date", "Expiry Date",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        short_url = row.get("short_url") or row.get("folder_url", "")
        ws.append([
            row["folder_name"],
            row.get("share_status", "Shared"),
            short_url,
            row.get("encrypted_url", ""),
            row["password"],
            row["creation_date"],
            row.get("expiry_date"),
        ])
        # The short URL is both the display text AND the hyperlink target —
        # copy-pasting it works, clicking it works, and no SharePoint path is visible.
        if short_url:
            url_cell = ws.cell(row=ws.max_row, column=3)
            url_cell.hyperlink = short_url
            url_cell.font = Font(color="0563C1", underline="single")

    for column, width in {
        "A": 32, "B": 16, "C": 55, "D": 70, "E": 28, "F": 22, "G": 15,
    }.items():
        ws.column_dimensions[column].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# T-12: ReportStats dataclass
# ---------------------------------------------------------------------------


@dataclass
class ReportStats:
    """Accumulates metadata about a completed report run."""

    subfolders_found: int = 0
    rows_generated: int = 0
    uploaded_filename: str | None = None
    skipped_reason: str | None = None
    errors: list[str] = dc_field(default_factory=list)


# ---------------------------------------------------------------------------
# T-12: collect_subfolders
# ---------------------------------------------------------------------------


def collect_subfolders(page: "Page", source_folder: str) -> list[str]:  # type: ignore[name-defined]
    """Navigate to source_folder and return immediate subfolder names.

    Args:
        page: Authenticated Playwright page.
        source_folder: Relative OneDrive path to enumerate.

    Returns:
        List of subfolder names (strings) found at the immediate level.

    Note:
        list_items() defaults to exhaustive=True, which scrolls the listing
        until the row count stabilizes before reading the DOM — this covers
        the OneDrive DOM virtualization case that used to truncate large
        folders (previously a documented >= 50 items limitation here).
    """
    from playwright.sync_api import Page  # local import keeps module loadable without playwright

    from onedrive_rpa.rpa._navigation import navigate_to_folder, list_items, FolderNotFoundError

    try:
        navigate_to_folder(page, source_folder)
    except FolderNotFoundError:
        raise ReportError(f"source_folder not found: {source_folder}")

    # navigate_to_folder already waits for folder_row to be attached.
    # list_items repeats that wait internally — no additional delay needed.
    # networkidle is intentionally avoided: SharePoint continuously fires
    # background telemetry requests and never truly reaches idle, so waiting
    # for it burns the full timeout (15 s) on every call.
    items = list_items(page)
    folders = [item.name for item in items if item.is_folder]

    return folders


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


def _click_first_matching(page: "Page", selectors: list[str], timeout_ms: int) -> bool:  # type: ignore[name-defined]
    """Try each selector in order and click the first one that resolves. Returns True on success."""
    for sel in selectors:
        sel = sel.strip()
        if not sel:
            continue
        try:
            page.click(sel, timeout=timeout_ms)
            return True
        except Exception:
            continue
    return False


# ---------------------------------------------------------------------------
# T-12: upload_report
# ---------------------------------------------------------------------------


def upload_report(
    page: "Page",  # type: ignore[name-defined]
    excel_bytes: "BytesIO | bytes",
    destination_folder: str,
    filename: str,
) -> None:
    """Upload an Excel file to destination_folder in OneDrive via Playwright.

    Tries Shape A first (hidden file input via set_input_files). Falls back to
    Shape B (toolbar Upload button + file chooser) if Shape A is unavailable.

    Args:
        page: Authenticated Playwright page.
        excel_bytes: Either a BytesIO object or raw bytes.
        destination_folder: Relative OneDrive path for the upload destination.
        filename: The filename the uploaded file should have in OneDrive.

    Raises:
        ReportError: If the upload confirmation selector never appears.
    """
    from onedrive_rpa.rpa._navigation import navigate_to_folder, FolderNotFoundError
    from onedrive_rpa.config import SELECTORS, ACTION_TIMEOUT_MS, UPLOAD_TIMEOUT_MS

    try:
        navigate_to_folder(page, destination_folder)
    except FolderNotFoundError:
        raise ReportError(f"destination_folder not found: {destination_folder}")

    # Resolve bytes from BytesIO if needed
    if hasattr(excel_bytes, "read"):
        data: bytes = excel_bytes.read()
    else:
        data = excel_bytes  # type: ignore[assignment]

    tmp = tempfile.NamedTemporaryFile(
        prefix="onedrive_report_", suffix=".xlsx", delete=False
    )
    target_path: Path | None = None
    try:
        tmp.write(data)
        tmp.close()

        # Rename temp file to the intended filename so SharePoint shows the right name
        target_path = Path(tmp.name).parent / filename
        os.replace(tmp.name, target_path)

        # Shape A is intentionally SKIPPED here.
        # A stale input[type='file'] left in the DOM from the deletion phase
        # belongs to a different folder context — using it would upload to the
        # wrong folder.  We always go through Shape B so the navigation to
        # destination_folder is the authoritative upload context.

        # Shape B: click "Crear o cargar" → "Carga de archivos".
        # SharePoint triggers a native OS file chooser when the submenu item
        # is clicked.  We MUST register expect_file_chooser BEFORE the click
        # so Playwright intercepts the dialog before it blocks execution.
        _click_first_matching(page, SELECTORS["toolbar_upload"].split(", "), ACTION_TIMEOUT_MS)

        # Wait for the submenu item itself to become visible — more reliable
        # than waiting for a generic [role='menu'] container that may not match.
        # Falls back to a fixed 800 ms pause if the selector doesn't resolve.
        submenu_selectors = ", ".join([
            "[data-automationid='uploadFilesCommand']",
            "[data-automationid='uploadFileCommand']",
            "[role='menuitem']:has-text('Carga de archivos')",
            "button:has-text('Carga de archivos')",
        ])
        try:
            page.wait_for_selector(submenu_selectors, timeout=5_000, state="visible")
        except Exception:
            page.wait_for_timeout(800)

        try:
            with page.expect_file_chooser(timeout=15_000) as fc_info:
                _click_first_matching(
                    page, SELECTORS["upload_files_menuitem"].split(", "), 3_000
                )
            fc_info.value.set_files(str(target_path))
        except Exception:
            # Fallback: some SP versions inject the input without OS dialog
            hidden_input = page.locator(SELECTORS["upload_file_input"]).first
            hidden_input.wait_for(state="attached", timeout=10_000)
            hidden_input.set_input_files(str(target_path))

        # Wait for the uploaded file row to appear as confirmation.
        # Try two selectors: heroField (SharePoint item name) and a broad text match.
        # FieldRenderer-name is unreliable across tenants/versions.
        confirmed = False
        for confirm_sel in [
            f"[data-id='heroField']:has-text('{filename}')",
            f"text={filename}",
        ]:
            try:
                page.wait_for_selector(confirm_sel, timeout=UPLOAD_TIMEOUT_MS, state="visible")
                confirmed = True
                break
            except Exception:
                pass
        if not confirmed:
            raise ReportError(
                f"Upload confirmation not found for '{filename}' within timeout"
            )

    finally:
        for p in [Path(tmp.name), target_path]:
            if p is not None and Path(p).exists():
                try:
                    os.unlink(p)
                except OSError:
                    pass


# ---------------------------------------------------------------------------
# T-12: run_report
# ---------------------------------------------------------------------------


def run_report(
    page: "Page",  # type: ignore[name-defined]
    source_folder: str,
    destination_folder: str,
    *,
    callbacks=None,
    passwords: dict[str, str] | None = None,
    share_urls: dict[str, str] | None = None,
    share_statuses: dict[str, str] | None = None,
    folder_paths: list[str] | None = None,
    expiry_date: "datetime | None" = None,
) -> ReportStats:
    """Orchestrate collect → build rows → write Excel → upload.

    Non-fatal: all exceptions are caught and surfaced via ReportStats.errors
    so that the caller (main.py) can continue to the summary step.

    Args:
        page: Authenticated Playwright page.
        source_folder: Folder to enumerate immediate subfolders from (used when
                       ``folder_paths`` is not provided). Also used to navigate
                       to the destination for upload.
        destination_folder: Folder where the Excel report is uploaded.
        callbacks: Optional RPACallbacks-compatible object. Calls are guarded
                   with hasattr so any subset of callbacks is acceptable.
        passwords: Optional pre-generated password map (keyed by folder base
                   name) forwarded to ``build_report_rows()``.  When provided,
                   ensures the password in the report matches the one set on the
                   sharing link.  When ``None``, each row generates a fresh
                    password (backward compatible).
        share_statuses: Optional map of folder base name to its sharing outcome.
                        Failed, skipped, and not-shared folders are emitted with
                        blank link and password fields.
        folder_paths: Optional list of full folder paths from the ``clean`` list
                      (e.g. ``["Camion/Plano/Bz8yy", "Camion/ADMIN/Bz2rr"]``).
                      When provided, the report rows are derived directly from
                      these paths (leaf name as display name, full path for URL
                      construction) instead of enumerating ``source_folder`` via
                      the DOM. This ensures the report shows the same specific
                      folders that were cleaned and shared.

    Returns:
        ReportStats with counts, filename, and any errors encountered.
    """
    stats = ReportStats()

    try:
        if callbacks and hasattr(callbacks, "on_report_start"):
            callbacks.on_report_start(source_folder, destination_folder)

        if folder_paths:
            # Derive rows directly from the clean paths — leaf name as display
            # name, full path for URL construction. Avoids DOM navigation and
            # ensures report rows match exactly what was cleaned and shared.
            leaf_names = [fp.rstrip("/").rsplit("/", 1)[-1] for fp in folder_paths]
            subfolders = leaf_names
            full_paths_for_rows: list[str] | None = folder_paths
        else:
            subfolders = collect_subfolders(page, source_folder)
            full_paths_for_rows = None

        # Exclude the destination folder itself from the report when it lives
        # directly inside source_folder (e.g. source="pruebas", dest="pruebas/registro"
        # → exclude "registro" so the archive folder doesn't appear as a data row).
        source_prefix = source_folder.rstrip("/") + "/"
        if destination_folder.startswith(source_prefix):
            dest_child = destination_folder[len(source_prefix):].split("/")[0]
            if full_paths_for_rows:
                pairs = [
                    (n, p) for n, p in zip(subfolders, full_paths_for_rows)
                    if n != dest_child
                ]
                subfolders, full_paths_for_rows = (
                    [n for n, _ in pairs], [p for _, p in pairs]
                )
            else:
                subfolders = [f for f in subfolders if f != dest_child]

        stats.subfolders_found = len(subfolders)

        if callbacks and hasattr(callbacks, "on_report_subfolders"):
            callbacks.on_report_subfolders(len(subfolders))

        rows = build_report_rows(
            subfolders, source_folder=source_folder, fernet=config.FERNET,
            passwords=passwords, share_urls=share_urls, share_statuses=share_statuses,
            full_paths=full_paths_for_rows, expiry_date=expiry_date,
        )
        stats.rows_generated = len(rows)

        excel_buf = write_excel(rows)
        filename = build_report_filename()

        upload_report(page, excel_buf, destination_folder, filename)
        stats.uploaded_filename = filename

        if callbacks and hasattr(callbacks, "on_report_uploaded"):
            callbacks.on_report_uploaded(filename)

        logger.info(
            "REPORT_END | rows={n} | uploaded={f} | destination={d}",
            n=len(rows),
            f=filename,
            d=destination_folder,
        )

    except ReportError as exc:
        stats.errors.append(str(exc))
        logger.error("REPORT_ERROR | reason={r}", r=str(exc))
        if callbacks and hasattr(callbacks, "on_report_error"):
            callbacks.on_report_error(str(exc))

    except Exception as exc:
        stats.errors.append(str(exc))
        logger.error("REPORT_ERROR | unexpected | reason={r}", r=str(exc))
        if callbacks and hasattr(callbacks, "on_report_error"):
            callbacks.on_report_error(str(exc))

    return stats
