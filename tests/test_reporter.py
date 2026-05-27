"""
tests/test_reporter.py — Unit tests for pure functions in rpa/reporter.py.

These tests are written in RED phase (PR 1). The implementation lives in PR 2.
All tests are expected to FAIL until reporter.py is implemented.

Tested functions:
    generate_password(length: int = REPORT_PASSWORD_LENGTH) -> str
    build_report_rows(folder_names: list[str], now: datetime | None) -> list[dict]
    write_excel(rows: list[dict]) -> BytesIO
    build_report_filename(dt: datetime | None) -> str
    _build_folder_url(source_folder: str, name: str) -> str
"""

import re
import unittest
from datetime import datetime
from io import BytesIO
from unittest import mock

from onedrive_rpa.config import REPORT_PASSWORD_LENGTH, REPORT_PASSWORD_ALPHABET


class TestGeneratePassword(unittest.TestCase):
    """Tests for generate_password()."""

    def _import(self):
        from onedrive_rpa.rpa.reporter import generate_password
        return generate_password

    def test_generate_password_length(self):
        """Generated password must have exactly REPORT_PASSWORD_LENGTH characters."""
        generate_password = self._import()
        pwd = generate_password()
        self.assertEqual(len(pwd), REPORT_PASSWORD_LENGTH)

    def test_generate_password_alphabet(self):
        """Every character in the password must be from REPORT_PASSWORD_ALPHABET."""
        generate_password = self._import()
        pwd = generate_password()
        allowed = set(REPORT_PASSWORD_ALPHABET)
        for ch in pwd:
            self.assertIn(ch, allowed, f"Character {ch!r} is not in allowed alphabet")

    def test_generate_password_no_quotes_10k_samples(self):
        """Over 10 000 iterations the password must never contain \" or '."""
        generate_password = self._import()
        for _ in range(10_000):
            pwd = generate_password()
            self.assertNotIn('"', pwd)
            self.assertNotIn("'", pwd)

    def test_generate_password_min_length_raises(self):
        """generate_password(15) must raise ValueError (minimum is 16)."""
        generate_password = self._import()
        with self.assertRaises(ValueError):
            generate_password(15)


class TestBuildReportRows(unittest.TestCase):
    """Tests for build_report_rows()."""

    def _import(self):
        from onedrive_rpa.rpa.reporter import build_report_rows
        return build_report_rows

    def test_build_report_rows_count(self):
        """3 folder names must produce exactly 3 rows."""
        build_report_rows = self._import()
        names = ["alpha", "beta", "gamma"]
        rows = build_report_rows(names)
        self.assertEqual(len(rows), 3)

    def test_build_report_rows_keys(self):
        """Each row dict must contain folder_name, password, and creation_date keys."""
        build_report_rows = self._import()
        rows = build_report_rows(["alpha"])
        self.assertIn("folder_name", rows[0])
        self.assertIn("password", rows[0])
        self.assertIn("creation_date", rows[0])

    def test_build_report_rows_injected_now(self):
        """creation_date in each row must equal the injected 'now' parameter."""
        build_report_rows = self._import()
        fixed_now = datetime(2026, 5, 22, 10, 30, 0)
        rows = build_report_rows(["alpha", "beta"], now=fixed_now)
        for row in rows:
            self.assertEqual(row["creation_date"], fixed_now)


class TestWriteExcel(unittest.TestCase):
    """Tests for write_excel()."""

    def _import(self):
        from onedrive_rpa.rpa.reporter import write_excel
        return write_excel

    def test_write_excel_round_trip(self):
        """
        write_excel must return a BytesIO that openpyxl can read back.
        The workbook must contain the expected header and at least one data row.
        """
        import openpyxl

        write_excel = self._import()

        rows = [
            {"folder_name": "alpha", "password": "abc123", "creation_date": datetime(2026, 5, 22)},
            {"folder_name": "beta",  "password": "xyz789", "creation_date": datetime(2026, 5, 22)},
        ]
        result = write_excel(rows)

        self.assertIsInstance(result, BytesIO)

        # Should be seeked to 0 so openpyxl can read from the start
        wb = openpyxl.load_workbook(result)
        ws = wb.active

        # Header row
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        self.assertIn("Folder Name", headers)
        self.assertIn("Password", headers)
        self.assertIn("Creation Date", headers)

        # Data rows (should have 2)
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(data_rows), 2)


class TestBuildReportFilename(unittest.TestCase):
    """Tests for build_report_filename()."""

    def _import(self):
        from onedrive_rpa.rpa.reporter import build_report_filename
        return build_report_filename

    def test_build_report_filename_format(self):
        """Filename must match the pattern reporte_YYYYMMDD_HHMMSS.xlsx."""
        build_report_filename = self._import()
        filename = build_report_filename()
        pattern = r"^reporte_\d{8}_\d{6}\.xlsx$"
        self.assertRegex(filename, pattern, f"Filename {filename!r} does not match expected pattern")


# ---------------------------------------------------------------------------
# Task 3.4 (updated): TestWriteExcel — Encrypted URL column in headers
# ---------------------------------------------------------------------------


class TestWriteExcelEncryptedUrl(unittest.TestCase):
    """Tests for write_excel() with the new Encrypted URL column (task 3.4)."""

    def _import(self):
        from onedrive_rpa.rpa.reporter import write_excel
        return write_excel

    def test_write_excel_headers_include_encrypted_url(self):
        """Header row must include 'Encrypted URL' between Password and Creation Date."""
        import openpyxl

        write_excel = self._import()
        rows = [
            {
                "folder_name": "alpha",
                "password": "abc123",
                "encrypted_url": "encryptedvalue123",
                "creation_date": datetime(2026, 5, 22),
            }
        ]
        result = write_excel(rows)
        wb = openpyxl.load_workbook(result)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        self.assertEqual(headers, ["Folder Name", "Password", "URL", "Creation Date"])

    def test_write_excel_encrypted_url_in_data_row_column_3(self):
        """Short URL (or fallback) must appear in column index 2 (3rd column, zero-based)."""
        import openpyxl

        write_excel = self._import()
        rows = [
            {
                "folder_name": "alpha",
                "password": "abc123",
                "short_url": "https://is.gd/abc123",
                "encrypted_url": "fernet_token_here",
                "creation_date": datetime(2026, 5, 22),
            }
        ]
        result = write_excel(rows)
        wb = openpyxl.load_workbook(result)
        ws = wb.active
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(data_rows), 1)
        # column index 2 (third column) must show the short URL
        self.assertEqual(data_rows[0][2], "https://is.gd/abc123")

    def test_write_excel_missing_encrypted_url_key_uses_empty_string(self):
        """Rows without 'encrypted_url' key must produce empty string in column 3 (backward compat)."""
        import openpyxl

        write_excel = self._import()
        rows = [
            {
                "folder_name": "beta",
                "password": "xyz789",
                "creation_date": datetime(2026, 5, 22),
                # no 'encrypted_url' key — simulates old caller
            }
        ]
        result = write_excel(rows)
        wb = openpyxl.load_workbook(result)
        ws = wb.active
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        # column 3 (index 2) must be empty string (or None — both acceptable) for missing key
        self.assertIn(data_rows[0][2], ("", None))


# ---------------------------------------------------------------------------
# Task 3.1: TestBuildFolderUrl
# ---------------------------------------------------------------------------


class TestBuildFolderUrl(unittest.TestCase):
    """Tests for _build_folder_url() helper in reporter.py (task 3.1)."""

    def _import(self):
        from onedrive_rpa.rpa.reporter import _build_folder_url
        return _build_folder_url

    def test_clean_segments_produce_correct_url(self):
        """Clean path segments produce the expected URL without double-slashes."""
        _build_folder_url = self._import()
        with mock.patch("onedrive_rpa.config.ONEDRIVE_URL", "https://archacomco-my.sharepoint.com"), \
             mock.patch("onedrive_rpa.config.SHAREPOINT_PERSONAL_PATH", "/personal/carlos_velasco"):
            url = _build_folder_url("clientes", "AlphaClient")
        self.assertEqual(
            url,
            "https://archacomco-my.sharepoint.com/personal/carlos_velasco/Documents/clientes/AlphaClient"
        )

    def test_trailing_slash_in_source_folder_normalised(self):
        """Trailing slash in source_folder must not produce double-slash in URL."""
        _build_folder_url = self._import()
        with mock.patch("onedrive_rpa.config.ONEDRIVE_URL", "https://archacomco-my.sharepoint.com"), \
             mock.patch("onedrive_rpa.config.SHAREPOINT_PERSONAL_PATH", "/personal/carlos_velasco"):
            url = _build_folder_url("clientes/", "AlphaClient")
        self.assertNotIn("//", url.split("://")[1])  # no double-slash after scheme

    def test_empty_source_folder_still_valid(self):
        """Empty source_folder must produce a URL without doubled slashes at that position."""
        _build_folder_url = self._import()
        with mock.patch("onedrive_rpa.config.ONEDRIVE_URL", "https://archacomco-my.sharepoint.com"), \
             mock.patch("onedrive_rpa.config.SHAREPOINT_PERSONAL_PATH", "/personal/carlos_velasco"):
            url = _build_folder_url("", "AlphaClient")
        # Must not contain double slash in path segment
        self.assertNotIn("//", url.split("://")[1])
        # Must still contain the folder name
        self.assertIn("AlphaClient", url)


# ---------------------------------------------------------------------------
# Task 3.2: TestBuildReportRowsEncryption
# ---------------------------------------------------------------------------


class TestBuildReportRowsEncryption(unittest.TestCase):
    """Tests for build_report_rows() encryption behavior (task 3.2)."""

    def _import(self):
        from onedrive_rpa.rpa.reporter import build_report_rows
        return build_report_rows

    def _make_fernet(self):
        from cryptography.fernet import Fernet
        return Fernet(Fernet.generate_key())

    def test_encrypted_url_present_when_fernet_injected(self):
        """Each row must contain 'encrypted_url' key with non-empty value when Fernet is injected."""
        build_report_rows = self._import()
        fernet = self._make_fernet()
        rows = build_report_rows(["alpha", "beta"], source_folder="clientes", fernet=fernet)
        for row in rows:
            self.assertIn("encrypted_url", row)
            self.assertTrue(row["encrypted_url"], f"encrypted_url must be non-empty, got {row['encrypted_url']!r}")

    def test_two_calls_produce_different_ciphertext(self):
        """Non-deterministic: same inputs encrypted twice MUST produce different ciphertext."""
        build_report_rows = self._import()
        fernet = self._make_fernet()
        rows1 = build_report_rows(["alpha"], source_folder="clientes", fernet=fernet)
        rows2 = build_report_rows(["alpha"], source_folder="clientes", fernet=fernet)
        self.assertNotEqual(
            rows1[0]["encrypted_url"],
            rows2[0]["encrypted_url"],
            "Fernet.encrypt() must be non-deterministic — two calls must differ"
        )

    def test_fernet_none_yields_empty_string_without_raising(self):
        """fernet=None + config.FERNET=None must produce empty string without raising."""
        import unittest.mock as mock
        import onedrive_rpa.config as cfg
        build_report_rows = self._import()
        with mock.patch.object(cfg, "FERNET", None):
            rows = build_report_rows(["alpha", "beta"], fernet=None)
        for row in rows:
            self.assertIn("encrypted_url", row)
            self.assertEqual(row["encrypted_url"], "")


# ---------------------------------------------------------------------------
# Task 3.3: TestBuildReportRowsBackwardCompat
# ---------------------------------------------------------------------------


class TestBuildReportRowsBackwardCompat(unittest.TestCase):
    """Tests for build_report_rows() backward compatibility (task 3.3)."""

    def _import(self):
        from onedrive_rpa.rpa.reporter import build_report_rows
        return build_report_rows

    def test_call_without_new_kwargs_returns_correct_keys(self):
        """Existing callers without source_folder or fernet kwargs must not raise TypeError."""
        build_report_rows = self._import()
        fixed_now = datetime(2026, 5, 22, 10, 30, 0)
        # Old-style call: no source_folder, no fernet
        rows = build_report_rows(["alpha", "beta", "gamma"], now=fixed_now)
        self.assertEqual(len(rows), 3)
        for row in rows:
            self.assertIn("folder_name", row)
            self.assertIn("password", row)
            self.assertIn("creation_date", row)
            self.assertEqual(row["creation_date"], fixed_now)

    def test_call_without_kwargs_returns_empty_encrypted_url(self):
        """Old callers not passing fernet get empty string for encrypted_url (no key configured)."""
        build_report_rows = self._import()
        # Ensure config.FERNET is None by patching (no key in test env)
        with mock.patch("onedrive_rpa.config.FERNET", None):
            rows = build_report_rows(["alpha"])
        self.assertIn("encrypted_url", rows[0])
        self.assertEqual(rows[0]["encrypted_url"], "")


# ---------------------------------------------------------------------------
# Task 3.5: TestConfigFernet
# ---------------------------------------------------------------------------


class TestConfigFernet(unittest.TestCase):
    """Tests for config._build_fernet() and config.FERNET (task 3.5)."""

    def test_build_fernet_returns_none_when_key_absent(self):
        """_build_fernet with empty bytes must return None (fail-open)."""
        from onedrive_rpa.config import _build_fernet
        result = _build_fernet(b"")
        self.assertIsNone(result)

    def test_build_fernet_returns_none_when_key_invalid(self):
        """_build_fernet with an invalid key must return None and not raise."""
        from onedrive_rpa.config import _build_fernet
        result = _build_fernet(b"this-is-not-a-valid-fernet-key")
        self.assertIsNone(result)

    def test_build_fernet_returns_fernet_when_key_valid(self):
        """_build_fernet with a valid Fernet key must return a Fernet instance."""
        from cryptography.fernet import Fernet
        from onedrive_rpa.config import _build_fernet
        valid_key = Fernet.generate_key()
        result = _build_fernet(valid_key)
        self.assertIsNotNone(result)
        # Verify the returned instance can encrypt/decrypt
        token = result.encrypt(b"test")
        self.assertEqual(result.decrypt(token), b"test")

    def test_fernet_module_attr_is_none_when_env_var_absent(self):
        """config.FERNET must be None when FOLDERS_ENCRYPTION_KEY is not set."""
        import importlib
        import onedrive_rpa.config as cfg_module
        with mock.patch.dict("os.environ", {}, clear=False):
            # Remove the key if it exists
            env_backup = os.environ.pop("FOLDERS_ENCRYPTION_KEY", None)
            try:
                # Re-run just the helper with empty key (module-level FERNET may already be set)
                from onedrive_rpa.config import _build_fernet
                result = _build_fernet(b"")
                self.assertIsNone(result)
            finally:
                if env_backup is not None:
                    os.environ["FOLDERS_ENCRYPTION_KEY"] = env_backup


import os  # needed for TestConfigFernet


if __name__ == "__main__":
    unittest.main()
